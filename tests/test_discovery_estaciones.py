# -*- coding: utf-8 -*-
"""
Pruebas de discovery_estaciones.py.

Los XML y los CSV de prueba son inventados, con la misma forma exacta que
entrega la descarga de facturas de Syntage (se confirmó contra un zip real
de 17 clientes, que no se sube al repo por traer RFC y razón social de
clientes reales). Los zips se arman aquí mismo, en memoria y en un
directorio temporal: la prueba no depende de que exista ninguna descarga.

Se corre con:
    python tests/test_discovery_estaciones.py
"""

import io
import os
import sys
import tempfile
import zipfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import discovery_estaciones as de

fallas = []


def check(cond, msg):
    print(("  ok  " if cond else "FALLA ") + msg)
    if not cond:
        fallas.append(msg)


# ─────────────────────────────────────────────────────────────────────────────
# Insumos: CFDI en XML con y sin el complemento de combustible
# ─────────────────────────────────────────────────────────────────────────────
def xml_con_complemento(uuid, rfc_monedero, nombre_monedero, rfc_cliente,
                        nombre_cliente, fecha, cargos):
    """Un CFDI con el complemento 'Estado de Cuenta de Combustibles'. El
    SubTotal del comprobante es simbólico ($1) y el real vive en el
    complemento — así es en los CFDI reales de monedero."""
    conceptos = "".join(
        '<ecc12:ConceptoEstadoDeCuentaCombustible Identificador="%s" Fecha="%sT08:00:00" '
        'Rfc="%s" ClaveEstacion="%s" Cantidad="%s" TipoCombustible="01" '
        'NombreCombustible="Magna" FolioOperacion="%s" ValorUnitario="%s" '
        'Importe="%s"/>' % (c["id"], fecha, c["rfc_estacion"], c["clave"],
                            c["litros"], c["id"], c["unitario"], c["importe"])
        for c in cargos)
    subtotal = sum(c["importe"] for c in cargos)
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4" '
        'xmlns:ecc12="http://www.sat.gob.mx/EstadoDeCuentaCombustible12" '
        'xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital" '
        'Version="4.0" Fecha="%sT23:59:59" SubTotal="1.00" Total="1.00" TipoDeComprobante="I">'
        '<cfdi:Emisor Rfc="%s" Nombre="%s"/>'
        '<cfdi:Receptor Rfc="%s" Nombre="%s"/>'
        '<cfdi:Complemento>'
        '<ecc12:EstadoDeCuentaCombustible Version="1.2" TipoOperacion="Tarjeta" '
        'NumeroDeCuenta="F000001" SubTotal="%.2f" Total="%.2f">'
        '<ecc12:Conceptos>%s</ecc12:Conceptos>'
        '</ecc12:EstadoDeCuentaCombustible>'
        '<tfd:TimbreFiscalDigital UUID="%s"/>'
        '</cfdi:Complemento>'
        '</cfdi:Comprobante>' % (fecha, rfc_monedero, nombre_monedero, rfc_cliente,
                                 nombre_cliente, subtotal, subtotal * 1.16,
                                 conceptos, uuid))


def xml_descuadrado(uuid, rfc_monedero, rfc_cliente, fecha):
    """Mismo complemento, pero el subtotal que declara el monedero no
    coincide con la suma de sus cargos: se leyó mal o llegó incompleto."""
    x = xml_con_complemento(uuid, rfc_monedero, "MONEDERO FICTICIO", rfc_cliente,
                            "CLIENTE FICTICIO", fecha,
                            [{"id": "1", "rfc_estacion": "EST010101AA1", "clave": "E1",
                              "litros": "100.00", "unitario": "20.00", "importe": 2000.0}])
    return x.replace('SubTotal="2000.00"', 'SubTotal="9999.00"')


def xml_cargos(uuid, rfc_monedero, nombre_monedero, rfc_cliente, fecha, conceptos):
    """Una factura de cargos: mismo emisor, sin complemento de combustible.

    `conceptos` es una lista de (descripción, importe). Las descripciones
    reales varían por monedero — se confirmó contra un zip real de 17
    clientes: TOKA factura "COMISION", Pluxee "CARGO DE COMISION",
    Efectivale "Cargo Administrativo", y además hay cargos que NO son
    comisión ("PLASTICOS", "MAS DESPENSA TARJETAS REPOSICIONES")."""
    cuerpo = "".join(
        '<cfdi:Concepto Descripcion="%s" Importe="%.2f"/>' % (d, m)
        for d, m in conceptos)
    subtotal = sum(m for _, m in conceptos)
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4" '
        'xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital" '
        'Version="4.0" Fecha="%sT12:00:00" SubTotal="%.2f" Total="%.2f" TipoDeComprobante="I">'
        '<cfdi:Emisor Rfc="%s" Nombre="%s"/>'
        '<cfdi:Receptor Rfc="%s" Nombre="CLIENTE FICTICIO"/>'
        '<cfdi:Conceptos>%s</cfdi:Conceptos>'
        '<cfdi:Complemento><tfd:TimbreFiscalDigital UUID="%s"/></cfdi:Complemento>'
        '</cfdi:Comprobante>' % (fecha, subtotal, subtotal * 1.16, rfc_monedero,
                                 nombre_monedero, rfc_cliente, cuerpo, uuid))


CSV_ENCABEZADO = ("uuid,type,issuedAt,subtotal,total,issuerRfc,issuerName,"
                  "receiverRfc,receiverName\n")


def csv_facturas(filas):
    return CSV_ENCABEZADO + "".join(
        "%s,%s,%s,%s,%s,%s,%s,%s,%s\n" % f for f in filas)


def zip_cliente(archivos):
    """Un zip de descarga de Syntage: los archivos van bajo 'files/'."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        for nombre, contenido in archivos.items():
            z.writestr("files/" + nombre, contenido)
    return buf.getvalue()


def zip_completo(por_cliente):
    """El zip exterior: una carpeta por RFC de cliente, y dentro los zips de
    cada descarga. Es la forma exacta de 'Facturas completas.zip'."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        for rfc, zips in por_cliente.items():
            for i, contenido in enumerate(zips):
                z.writestr("Facturas completas/%s/descarga-%d-invoices.zip" % (rfc, i),
                           contenido)
    return buf.getvalue()


def escribir_zip(contenido):
    ruta = os.path.join(tempfile.mkdtemp(prefix="_prueba_discovery_"), "facturas.zip")
    with open(ruta, "wb") as fh:
        fh.write(contenido)
    return ruta


# ─────────────────────────────────────────────────────────────────────────────
# Escenario principal: dos clientes, dos monederos, tres estaciones
# ─────────────────────────────────────────────────────────────────────────────
# ALFA carga en EST-1 y EST-2 con el monedero MON-A.
# BETA carga en EST-1 (la misma que ALFA) con el monedero MON-B.
# EST-1 es la estación interesante: menos importe que EST-2, pero DOS clientes.
XML_ALFA_MARZO = xml_con_complemento(
    "AAAA0001-0000-0000-0000-000000000001", "MOA010101AA1", "MONEDERO ALFA",
    "ALF020202BB2", "ALFA INDUSTRIAL", "2026-03",
    [{"id": "1", "rfc_estacion": "EST030303CC3", "clave": "E1",
      "litros": "100.00", "unitario": "20.00", "importe": 2000.0},
     {"id": "2", "rfc_estacion": "EST040404DD4", "clave": "E2",
      "litros": "300.00", "unitario": "20.00", "importe": 6000.0}])

XML_ALFA_ABRIL = xml_con_complemento(
    "AAAA0002-0000-0000-0000-000000000002", "MOA010101AA1", "MONEDERO ALFA",
    "ALF020202BB2", "ALFA INDUSTRIAL", "2026-04",
    [{"id": "3", "rfc_estacion": "EST030303CC3", "clave": "E1",
      "litros": "100.00", "unitario": "20.00", "importe": 2000.0},
     # EST-5 empata en importe con EST-1 ($5,000) pero la usa un solo
     # cliente: sirve para probar el desempate del ranking de estaciones.
     {"id": "5", "rfc_estacion": "EST050505JJ5", "clave": "E5",
      "litros": "250.00", "unitario": "20.00", "importe": 5000.0}])

# BETA carga en la MISMA estación que ALFA (mismo RFC) pero con otra clave de
# sucursal. El agregado tiene que juntarlas: en los datos reales la clave de
# estación viene en "0" en 244 de 397 casos, y un mismo RFC llega a tener 18
# claves distintas. Agrupar por (RFC, clave) partiría a la estación más
# grande en pedazos y la haría ver como si la usara un solo cliente.
XML_BETA_MARZO = xml_con_complemento(
    "BBBB0001-0000-0000-0000-000000000001", "MOB050505EE5", "MONEDERO BETA",
    "BET060606FF6", "BETA LOGISTICA", "2026-03",
    [{"id": "4", "rfc_estacion": "EST030303CC3", "clave": "E1B",
      "litros": "50.00", "unitario": "20.00", "importe": 1000.0}])

# Los cargos de ALFA. Marzo trae la comisión escrita como "Cargo
# Administrativo" y, en la misma factura, un cargo de plásticos que NO es
# comisión. Abril la trae con otra de las descripciones reales.
XML_ALFA_CARGOS_MARZO = xml_cargos(
    "AAAA0003-0000-0000-0000-000000000003", "MOA010101AA1", "MONEDERO ALFA",
    "ALF020202BB2", "2026-03",
    [("Cargo Administrativo", 400.0), ("PLASTICOS", 35.0)])

XML_ALFA_CARGOS_ABRIL = xml_cargos(
    "AAAA0004-0000-0000-0000-000000000004", "MOA010101AA1", "MONEDERO ALFA",
    "ALF020202BB2", "2026-04", [("CARGO DE COMISION", 700.0)])

# GAMA solo tiene CSV: se sabe qué monedero usa, no cuánto carga ni dónde.
CSV_GAMA = csv_facturas([
    ("CCCC0001-0000-0000-0000-000000000001", "E", "2026-05-10", "3.11", "3.11",
     "MOB050505EE5", "MONEDERO BETA", "GAM070707GG7", "GAMA TRANSPORTES"),
    ("CCCC0002-0000-0000-0000-000000000002", "E", "2026-06-10", "3.11", "3.11",
     "MOB050505EE5", "MONEDERO BETA", "GAM070707GG7", "GAMA TRANSPORTES"),
])

RUTA_ZIP = escribir_zip(zip_completo({
    # El primer zip de ALFA repite el XML de marzo: el mismo CFDI llega en
    # varias descargas, y sin deduplicar por UUID se contaría doble.
    "ALF020202BB2": [
        zip_cliente({"a.xml": XML_ALFA_MARZO, "a.csv": csv_facturas([])}),
        zip_cliente({"a.xml": XML_ALFA_MARZO, "b.xml": XML_ALFA_ABRIL,
                     "c.xml": XML_ALFA_CARGOS_MARZO,
                     "d.xml": XML_ALFA_CARGOS_ABRIL}),
    ],
    "BET060606FF6": [zip_cliente({"d.xml": XML_BETA_MARZO})],
    "GAM070707GG7": [zip_cliente({"e.csv": CSV_GAMA})],
}))

r = de.analizar(RUTA_ZIP)


# ── Deduplicar por folio fiscal ────────────────────────────────────────────
alfa = [c for c in r["clientes"] if c["rfc_cliente"] == "ALF020202BB2"][0]
check(alfa["importe"] == 15000.0,
      "ALFA suma $15,000 (marzo $8,000 + abril $7,000), no $23,000: el XML "
      "repetido en dos zips se contó una sola vez — importe=%r" % alfa["importe"])
check(alfa["meses"] == 2, "ALFA tiene 2 meses con detalle: %r" % alfa["meses"])
check(alfa["litros"] == 750.0, "ALFA suma 750 litros: %r" % alfa["litros"])
check(alfa["promedio_mensual"] == 7500.0,
      "ALFA promedia $7,500 al mes: %r" % alfa["promedio_mensual"])
check(alfa["razon_social"] == "ALFA INDUSTRIAL",
      "la razón social sale del propio CFDI, no del nombre del archivo: %r"
      % alfa["razon_social"])
check(alfa["monederos"] == ["MONEDERO ALFA"],
      "el monedero de ALFA: %r" % alfa["monederos"])


# ── Ranking de clientes: por importe, de mayor a menor ─────────────────────
con_detalle = [c["rfc_cliente"] for c in r["clientes"]]
check(con_detalle == ["ALF020202BB2", "BET060606FF6"],
      "los clientes vienen ordenados por importe, el más grande primero: %r"
      % con_detalle)


# ── Estaciones: por RFC, con clientes distintos ────────────────────────────
por_rfc = {e["rfc_estacion"]: e for e in r["estaciones"]}
est1 = por_rfc["EST030303CC3"]
est2 = por_rfc["EST040404DD4"]

check(est1["sucursales"] == 2,
      "EST-1 junta sus 2 claves de sucursal (E1 y E1B) en un solo renglón por "
      "RFC: %r" % est1["sucursales"])
check(sorted(est1["claves"]) == ["E1", "E1B"],
      "y nombra las claves que agrupó, para no perder el detalle: %r" % est1["claves"])
check(est1["clientes_distintos"] == 2,
      "EST-1 la usan 2 clientes distintos (ALFA y BETA): %r" % est1["clientes_distintos"])
check(est2["clientes_distintos"] == 1,
      "EST-2 la usa 1 solo cliente: %r" % est2["clientes_distintos"])
check(est1["importe"] == 5000.0,
      "EST-1 acumula $5,000 entre sus dos sucursales y sus dos clientes: %r"
      % est1["importe"])
check(est1["cargas"] == 3, "EST-1 tuvo 3 cargas: %r" % est1["cargas"])
check(est1["litros"] == 250.0, "EST-1 acumula 250 litros: %r" % est1["litros"])
check(est1["meses_activos"] == 2, "EST-1 estuvo activa 2 meses: %r" % est1["meses_activos"])
check(est1["promedio_mensual"] == 2500.0,
      "EST-1 promedia $2,500 al mes ($5,000 / 2 meses): %r" % est1["promedio_mensual"])
check(sorted(est1["clientes"]) == ["ALFA INDUSTRIAL", "BETA LOGISTICA"],
      "EST-1 nombra a sus clientes, para poder llamarles: %r" % est1["clientes"])

# El orden importa para el go-to-market. El criterio principal es el importe:
# el volumen que se puede redirigir a esa estación es el premio, y ordenar
# por clientes primero pondría una estación de $5,000 con dos clientes por
# encima de una de $500,000 con uno. Los clientes distintos son el DESEMPATE:
# entre dos estaciones del mismo tamaño, la que ya sirve a varios clientes se
# cierra una vez y beneficia a todos.
orden = [e["rfc_estacion"] for e in r["estaciones"]]
check(orden[0] == "EST040404DD4",
      "manda el importe: EST-2 ($6,000) va antes que EST-1 ($5,000) aunque "
      "EST-1 tenga más clientes — %r" % orden)
check(orden.index("EST030303CC3") < orden.index("EST050505JJ5"),
      "empatadas en $5,000, la de 2 clientes va antes que la de 1: %r" % orden)


# ── Comisión: las tres descripciones reales, y lo que no es comisión ──────
# Cada monedero nombra su comisión distinto. Se confirmó contra datos reales:
# TOKA factura "COMISION", Pluxee "CARGO DE COMISION", Efectivale "Cargo
# Administrativo". Buscar una sola de esas descripciones —lo que hace hoy
# estaciones_monedero.comision_candidatas() contra la API— encuentra casi nada.
check(alfa["comision"] == 1100.0,
      "la comisión de ALFA suma $1,100: reconoce 'Cargo Administrativo' ($400) "
      "y 'CARGO DE COMISION' ($700) como la misma cosa — %r" % alfa["comision"])
check(abs(alfa["comision_porcentaje"] - 1100.0 / 15000.0) < 1e-9,
      "la comisión de ALFA es 7.33%% ($1,100 sobre los $15,000 cargados en los "
      "dos meses que tienen ambas facturas): %r" % alfa["comision_porcentaje"])

# "PLASTICOS" es reposición de tarjetas, no comisión. No se suma a la
# comisión, pero tampoco se tira: se reporta aparte para que quien lea el
# Excel decida, en vez de desaparecer sin dejar rastro.
otros = {(o["descripcion"], o["rfc_cliente"]): o for o in r["otros_cargos"]}
check(("PLASTICOS", "ALF020202BB2") in otros,
      "el cargo que no es comisión se reporta aparte, no se descarta: %r"
      % list(otros))
check(otros[("PLASTICOS", "ALF020202BB2")]["importe"] == 35.0,
      "el importe del cargo ajeno: %r" % otros[("PLASTICOS", "ALF020202BB2")]["importe"])
check(all(o["descripcion"] != "PLASTICOS" or o["importe"] == 35.0
          for o in r["otros_cargos"]),
      "los $35 de plásticos no se colaron a la comisión de nadie")

beta = [c for c in r["clientes"] if c["rfc_cliente"] == "BET060606FF6"][0]
check(beta["comision"] == 0.0, "BETA no tiene factura de comisión: %r" % beta["comision"])
check(beta["comision_porcentaje"] is None,
      "sin factura de comisión el porcentaje es None, no cero: un cero diría "
      "que el monedero no le cobra nada — %r" % beta["comision_porcentaje"])


# ── Monederos ──────────────────────────────────────────────────────────────
mons = {m["rfc_monedero"]: m for m in r["monederos"]}
check(mons["MOA010101AA1"]["clientes"] == 1,
      "MONEDERO ALFA tiene 1 cliente: %r" % mons["MOA010101AA1"]["clientes"])
check(mons["MOB050505EE5"]["clientes"] == 2,
      "MONEDERO BETA tiene 2 clientes (BETA con XML y GAMA solo con CSV): %r"
      % mons["MOB050505EE5"]["clientes"])
check(abs(mons["MOA010101AA1"]["comision_porcentaje"] - 1100.0 / 15000.0) < 1e-9,
      "el %% de comisión de MONEDERO ALFA: %r"
      % mons["MOA010101AA1"]["comision_porcentaje"])
check(mons["MOB050505EE5"]["comision_porcentaje"] is None,
      "un monedero del que no se conoce comisión reporta None, no cero: %r"
      % mons["MOB050505EE5"]["comision_porcentaje"])


# ── Clientes sin detalle de estación: el CSV como única fuente ────────────
sin_detalle = {c["rfc_cliente"]: c for c in r["sin_detalle"]}
check("GAM070707GG7" in sin_detalle,
      "GAMA aparece en 'sin detalle': tiene CSV pero ningún CFDI descargado — %r"
      % list(sin_detalle))
check(sin_detalle["GAM070707GG7"]["monederos"] == ["MONEDERO BETA"],
      "de GAMA se sabe qué monedero usa, aunque no cuánto carga: %r"
      % sin_detalle["GAM070707GG7"]["monederos"])
check(sin_detalle["GAM070707GG7"]["razon_social"] == "GAMA TRANSPORTES",
      "la razón social de GAMA sale del CSV: %r"
      % sin_detalle["GAM070707GG7"]["razon_social"])
check(sin_detalle["GAM070707GG7"]["facturas"] == 2,
      "GAMA tiene 2 facturas listadas en su CSV: %r"
      % sin_detalle["GAM070707GG7"]["facturas"])
check("ALF020202BB2" not in sin_detalle,
      "un cliente que sí tiene detalle no se repite en 'sin detalle'")


# ── Lo que no cuadra se marca, no se promedia ─────────────────────────────
RUTA_DESCUADRE = escribir_zip(zip_completo({
    "DES080808HH8": [zip_cliente({
        "ok.xml": xml_con_complemento(
            "DDDD0001-0000-0000-0000-000000000001", "MOA010101AA1", "MONEDERO ALFA",
            "DES080808HH8", "DELTA SA", "2026-03",
            [{"id": "1", "rfc_estacion": "EST090909II9", "clave": "E9",
              "litros": "10.00", "unitario": "20.00", "importe": 200.0}]),
        "malo.xml": xml_descuadrado("DDDD0002-0000-0000-0000-000000000002",
                                    "MOA010101AA1", "DES080808HH8", "2026-04"),
    })],
}))
rd = de.analizar(RUTA_DESCUADRE)
delta = [c for c in rd["clientes"] if c["rfc_cliente"] == "DES080808HH8"][0]
check(delta["importe"] == 200.0,
      "el CFDI descuadrado no entra al importe del cliente: %r" % delta["importe"])
check(len(rd["sospechosos"]) == 1,
      "el CFDI descuadrado queda en sospechosos: %r" % rd["sospechosos"])
check("DDDD0002" in rd["sospechosos"][0]["folio_fiscal"],
      "el sospechoso se identifica por su folio fiscal: %r" % rd["sospechosos"][0])
check("EST090909II9" in {e["rfc_estacion"] for e in rd["estaciones"]},
      "la estación del CFDI que sí cuadró sigue contando")


# ── Un zip vacío no truena ────────────────────────────────────────────────
vacio = de.analizar(escribir_zip(zip_completo({})))
check(vacio["clientes"] == [] and vacio["estaciones"] == [] and vacio["monederos"] == [],
      "un zip sin ninguna factura regresa forma vacía, no truena: %r" % vacio)


# ── Leer una carpeta, no solo un zip ──────────────────────────────────────
# La descarga de Syntage a veces ya viene extraída en descargas/monederos/.
carpeta = tempfile.mkdtemp(prefix="_prueba_discovery_carpeta_")
sub = os.path.join(carpeta, "BET060606FF6", "files", "invoice.cfdi.xml")
os.makedirs(sub)
with open(os.path.join(sub, "d.xml"), "w", encoding="utf-8") as fh:
    fh.write(XML_BETA_MARZO)
rc = de.analizar(carpeta)
check([c["rfc_cliente"] for c in rc["clientes"]] == ["BET060606FF6"],
      "también lee una carpeta ya extraída, no solo un zip: %r"
      % [c["rfc_cliente"] for c in rc["clientes"]])


# ── Zip anidado: la descarga de Syntage a veces trae un zip dentro del zip ─
RUTA_ANIDADA = escribir_zip(zip_completo({
    "BET060606FF6": [zip_cliente({"interior.zip": zip_cliente({"d.xml": XML_BETA_MARZO})})],
}))
ra = de.analizar(RUTA_ANIDADA)
check([c["rfc_cliente"] for c in ra["clientes"]] == ["BET060606FF6"],
      "baja al zip anidado en vez de darlo por perdido: %r"
      % [c["rfc_cliente"] for c in ra["clientes"]])


# ── CFDI 3.3: otro namespace, mismos datos ────────────────────────────────
# En la descarga real de 17 clientes, 6 CFDI son versión 3.3 (namespace
# cfd/3 en vez de cfd/4), 2 de ellos con complemento de combustible. El
# complemento vive en el mismo namespace en ambas versiones, así que un
# lector que solo entienda cfd/4 sí encuentra los cargos pero no el RFC del
# cliente — y un cargo que no se puede atribuir a nadie es peor que uno que
# no se leyó.
XML_33 = (XML_BETA_MARZO
          .replace("http://www.sat.gob.mx/cfd/4", "http://www.sat.gob.mx/cfd/3")
          .replace('Version="4.0"', 'Version="3.3"')
          .replace("BBBB0001", "3333BBBB"))
r33 = de.analizar(escribir_zip(zip_completo({
    "BET060606FF6": [zip_cliente({"v33.xml": XML_33})]})))
check([c["rfc_cliente"] for c in r33["clientes"]] == ["BET060606FF6"],
      "un CFDI 3.3 se atribuye a su cliente igual que uno 4.0: %r"
      % [c["rfc_cliente"] for c in r33["clientes"]])
check(r33["clientes"][0]["importe"] == 1000.0,
      "el importe de un CFDI 3.3: %r" % r33["clientes"][0]["importe"])
check(r33["clientes"][0]["razon_social"] == "BETA LOGISTICA",
      "la razón social de un CFDI 3.3: %r" % r33["clientes"][0]["razon_social"])

# Y la comisión de un 3.3 también, que vive en cfdi:Concepto — el nodo que
# sí cambia de namespace entre versiones.
XML_33_COMISION = (XML_ALFA_CARGOS_MARZO
                   .replace("http://www.sat.gob.mx/cfd/4", "http://www.sat.gob.mx/cfd/3")
                   .replace('Version="4.0"', 'Version="3.3"')
                   .replace("AAAA0003", "3333AAAA"))
r33c = de.analizar(escribir_zip(zip_completo({
    "ALF020202BB2": [zip_cliente({"a.xml": XML_ALFA_MARZO,
                                  "c33.xml": XML_33_COMISION})]})))
check(r33c["clientes"][0]["comision"] == 400.0,
      "la comisión de un CFDI 3.3 se lee igual: %r" % r33c["clientes"][0]["comision"])


# ── Un CFDI sin RFC de cliente no se atribuye a nadie ─────────────────────
# Si por cualquier razón no se puede saber de quién es una factura, sus
# cargos NO pueden entrar al agregado: inflarían un ranking sin dueño. Se
# marca y se deja fuera, mismo criterio que un descuadre.
XML_SIN_RECEPTOR = XML_BETA_MARZO.replace(
    '<cfdi:Receptor Rfc="BET060606FF6" Nombre="BETA LOGISTICA"/>', "")
rsr = de.analizar(escribir_zip(zip_completo({
    "BET060606FF6": [zip_cliente({"huerfano.xml": XML_SIN_RECEPTOR})]})))
check(rsr["clientes"] == [],
      "un CFDI sin receptor no crea un cliente fantasma: %r" % rsr["clientes"])
check(rsr["estaciones"] == [],
      "sus cargos tampoco entran al ranking de estaciones: %r" % rsr["estaciones"])
check(len(rsr["sospechosos"]) == 1,
      "queda marcado como sospechoso, no desaparece: %r" % rsr["sospechosos"])


# ── El Excel se escribe de verdad ─────────────────────────────────────────
destino = os.path.join(tempfile.mkdtemp(prefix="_prueba_discovery_xlsx_"), "d.xlsx")
de.escribir_xlsx(r, destino)
check(os.path.exists(destino) and os.path.getsize(destino) > 0,
      "escribir_xlsx() deja un archivo con contenido en %s" % destino)

import openpyxl

libro = openpyxl.load_workbook(destino)
check(libro.sheetnames == ["Clientes", "Estaciones", "Monederos",
                           "Sin detalle de estación", "Otros cargos",
                           "Sospechosos"],
      "las seis hojas, en orden: %r" % libro.sheetnames)
hoja = libro["Estaciones"]
check(hoja.cell(row=1, column=1).value == "RFC estación",
      "la hoja de estaciones trae encabezados legibles: %r"
      % hoja.cell(row=1, column=1).value)
check(hoja.cell(row=2, column=1).value == "EST040404DD4",
      "el primer renglón de Estaciones es la de mayor importe: %r"
      % hoja.cell(row=2, column=1).value)


print()
if fallas:
    print("%d prueba(s) fallaron" % len(fallas))
    sys.exit(1)
print("Todas las pruebas pasaron.")
