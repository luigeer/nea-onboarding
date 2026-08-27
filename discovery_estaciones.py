# -*- coding: utf-8 -*-
"""
discovery_estaciones.py — ¿A qué clientes y a qué gasolineras atacar?
=============================================================================
Análisis hacia atrás, sobre TODOS los clientes a la vez, para decidir el
go-to-market del closed loop: quién carga más gasolina, en qué estaciones
(por RFC) se concentra ese volumen, qué monedero usa cada quien y cuánto le
cobran de comisión.

Se distingue de los dos módulos que ya existen por la unidad de análisis:
`monederos.py` y `estaciones_monedero.py` responden por UN cliente durante su
onboarding; aquí la pregunta es del portafolio completo.

Uso:
    python discovery_estaciones.py resumen "<ruta del zip o carpeta>"
    python discovery_estaciones.py xlsx    "<ruta del zip o carpeta>"
    python discovery_estaciones.py xlsx    "<ruta>" --comisiones

Sin `--comisiones` no se toca la red: todo sale de los archivos descargados.
Con `--comisiones` se le pide a Syntage lo único que no está en esos archivos
—cuánto le cobra el monedero a cada cliente— vía `comisiones_monedero.py`.

**Dos fuentes, cada una con lo que sí puede dar.** Se confirmó contra una
descarga real de 17 clientes:

- El **XML con complemento** de combustible es la única fuente del gasto real
  y del RFC de la estación. El SubTotal del comprobante es simbólico ($1); el
  monto verdadero vive en el complemento.
- El **CSV** de la descarga lista todas las facturas del cliente, pero con
  subtotales de $0 a $3: no sirve para gasto. Sí sirve para saber QUÉ monedero
  usa un cliente del que no se bajó ningún CFDI — sin él, esos clientes se
  verían como si no tuvieran monedero, y son prospectos igual de válidos.

**Por qué se deduplica por folio fiscal.** El mismo CFDI llega en varias
descargas y además extraído en `files/`. Sin deduplicar, el cliente con más
facturas —el más interesante— es justo el que más se infla.

**Por qué la comisión se busca por varias descripciones.** Cada monedero
nombra su comisión distinto: TOKA factura "COMISION", Pluxee "CARGO DE
COMISION", Efectivale "Cargo Administrativo". Buscar una sola —lo que hace
`estaciones_monedero.comision_candidatas()` contra la API— encuentra casi
nada. Y hay cargos del mismo emisor que NO son comisión ("PLASTICOS",
reposición de tarjetas): esos se reportan aparte en vez de sumarse a la
comisión o desaparecer sin dejar rastro.

**Por qué la comisión de verdad no se busca aquí, sino en Syntage.** Los CFDI
que este módulo lee son los estados de cuenta —los que el paso 1 manda
descargar—, y ahí la comisión no está por diseño: sobre 17 clientes y $2.47M
de combustible, los archivos descargados solo revelan $22.20 de comisión.
Tampoco está escondida dentro del estado de cuenta: la razón Total/SubTotal
del complemento va de 1.1435 a 1.1570 en los cinco monederos, o sea puro IVA.
La comisión llega en facturas APARTE del mismo emisor, que no se descargaron,
y por eso se pide a la API: ver `comisiones_monedero.py`. Las columnas
`RE_COMISION` de aquí son el respaldo para lo que sí aparezca en los
archivos, no la respuesta a "cuánto le cobran".
"""

import csv
import io
import os
import re
import zipfile
import xml.etree.ElementTree as ET

import estado_cuenta_monedero as ecm

RAIZ = os.path.dirname(os.path.abspath(__file__))

_NS = {"cfdi": "http://www.sat.gob.mx/cfd/4"}

# Las descripciones con las que los monederos facturan su comisión. Se
# compilaron de datos reales, no de la norma: el SAT no estandariza este
# concepto, así que la lista crece cuando aparezca un monedero nuevo.
RE_COMISION = re.compile(r"comisi[oó]n|cargo\s+administrativo", re.IGNORECASE)

# Cuántos niveles de zip anidado se abren. La descarga de Syntage a veces
# mete un zip dentro del zip del cliente, dentro del zip completo: tres.
PROFUNDIDAD_MAXIMA_ZIP = 4


# ─────────────────────────────────────────────────────────────────────────────
# Fuentes: recorrer un zip (con sus anidados) o una carpeta ya extraída
# ─────────────────────────────────────────────────────────────────────────────
def _fuentes_zip(datos, nombre, profundidad=0):
    if profundidad >= PROFUNDIDAD_MAXIMA_ZIP:
        return
    try:
        z = zipfile.ZipFile(io.BytesIO(datos))
    except zipfile.BadZipFile:
        return
    for miembro in z.namelist():
        if miembro.endswith("/"):
            continue
        ruta = "%s!%s" % (nombre, miembro)
        ext = os.path.splitext(miembro)[1].lower()
        if ext == ".zip":
            for f in _fuentes_zip(z.read(miembro), ruta, profundidad + 1):
                yield f
        elif ext in (".xml", ".csv"):
            yield ruta, ext, z.read(miembro)


def _fuentes(ruta):
    """(nombre legible, extensión, bytes) de cada XML y CSV, venga de un zip
    —con sus anidados— o de una carpeta ya extraída en disco."""
    if os.path.isdir(ruta):
        for carpeta, _, archivos in os.walk(ruta):
            for a in sorted(archivos):
                completa = os.path.join(carpeta, a)
                ext = os.path.splitext(a)[1].lower()
                if ext == ".zip":
                    with open(completa, "rb") as fh:
                        for f in _fuentes_zip(fh.read(), completa):
                            yield f
                elif ext in (".xml", ".csv"):
                    with open(completa, "rb") as fh:
                        yield completa, ext, fh.read()
        return
    with open(ruta, "rb") as fh:
        for f in _fuentes_zip(fh.read(), os.path.basename(ruta)):
            yield f


# ─────────────────────────────────────────────────────────────────────────────
# Leer un CFDI: complemento de combustible, o cargos del monedero
# ─────────────────────────────────────────────────────────────────────────────
def _por_nombre_local(raiz, nombre):
    """Todos los nodos con ese nombre, sin importar el namespace.

    Hace falta porque la descarga real trae CFDI 3.3 (namespace cfd/3) junto
    con 4.0 (cfd/4): 6 de 86 en el zip de 17 clientes. El complemento de
    combustible vive en el MISMO namespace en las dos versiones, así que un
    lector atado a cfd/4 sí encuentra los cargos pero no el RFC del cliente
    — y un cargo que no se puede atribuir a nadie infla un ranking sin
    dueño. Buscar por nombre local vale para las dos y para la que venga."""
    return [n for n in raiz.iter() if n.tag.rsplit("}", 1)[-1] == nombre]


def _uno(raiz, nombre):
    nodos = _por_nombre_local(raiz, nombre)
    return nodos[0] if nodos else None


def _identidad(raiz):
    emisor = _uno(raiz, "Emisor")
    receptor = _uno(raiz, "Receptor")
    return {
        "rfc_monedero": emisor.get("Rfc") if emisor is not None else None,
        "nombre_monedero": emisor.get("Nombre") if emisor is not None else None,
        "rfc_cliente": receptor.get("Rfc") if receptor is not None else None,
        "razon_social": receptor.get("Nombre") if receptor is not None else None,
    }


def _mes(raiz):
    """El mes del comprobante. A diferencia de
    `estaciones_monedero._mes_facturacion()`, aquí NO se corrige la zona
    horaria: el atributo Fecha de un CFDI ya viene en hora local del emisor
    (el SAT no admite offset), mientras que el `issuedAt` de la API de
    Syntage viene en UTC. Restarle seis horas a una fecha que ya es local
    movería al mes anterior justo los estados de cuenta emitidos a las
    23:59:59 del último día — los más comunes."""
    return (raiz.get("Fecha") or "")[:7] or None


def _cargos_de_monedero(raiz):
    """(comisión, otros) de un CFDI sin complemento de combustible. La
    comisión es lo que empata con RE_COMISION; 'otros' son los demás cargos
    con importe, que se reportan sin sumarse a la comisión."""
    comision = 0.0
    otros = []
    for c in _por_nombre_local(raiz, "Concepto"):
        desc = (c.get("Descripcion") or "").strip()
        try:
            importe = float(c.get("Importe") or 0)
        except ValueError:
            continue
        if importe <= 0:
            continue
        if RE_COMISION.search(desc):
            comision += importe
        else:
            otros.append({"descripcion": desc, "importe": importe})
    return comision, otros


def _leer(ruta):
    """Todo lo que se necesita de un CFDI, o None si no es uno legible."""
    try:
        raiz = ET.fromstring(ruta[2])
    except ET.ParseError:
        return None
    datos = ecm._parsear_complemento(raiz)
    folio = (datos["encabezado"] or {}).get("folio_fiscal")
    if not folio:
        return None
    registro = dict(_identidad(raiz))
    registro.update({
        "folio_fiscal": folio,
        "mes": _mes(raiz),
        "archivo": ruta[0],
        "resumen": datos["resumen"],
        "cargos": datos["cargos"],
    })
    if datos["resumen"] is None:
        registro["comision"], registro["otros_cargos"] = _cargos_de_monedero(raiz)
    else:
        registro["comision"], registro["otros_cargos"] = 0.0, []
    return registro


def _leer_csv(datos):
    """Las facturas que lista un CSV de la descarga: solo identidad y
    cadencia. Los subtotales de estos CSV son simbólicos ($0 a $3), así que
    deliberadamente no se lee ningún monto de aquí."""
    try:
        texto = datos.decode("utf-8")
    except UnicodeDecodeError:
        texto = datos.decode("latin-1", "replace")
    filas = []
    for f in csv.DictReader(io.StringIO(texto)):
        if not f.get("uuid"):
            continue
        filas.append({
            "folio_fiscal": f["uuid"],
            "rfc_cliente": (f.get("receiverRfc") or "").strip(),
            "razon_social": (f.get("receiverName") or "").strip(),
            "rfc_monedero": (f.get("issuerRfc") or "").strip(),
            "nombre_monedero": (f.get("issuerName") or "").strip(),
            "mes": (f.get("issuedAt") or "")[:7],
        })
    return filas


# ─────────────────────────────────────────────────────────────────────────────
# El análisis
# ─────────────────────────────────────────────────────────────────────────────
def _nombre_monedero(rfc, respaldo):
    """El nombre comercial del padrón del SAT si el RFC está ahí; si no, el
    nombre que trae el propio CFDI. El padrón dice 'Efecticard', el CFDI dice
    'EFECTIVALE': para hablar con un cliente sirve más el comercial."""
    try:
        import monederos
        m = monederos._PADRON_POR_RFC.get(rfc)
    except Exception:                                   # pragma: no cover
        m = None
    return (m or {}).get("nombre_comercial") or respaldo or rfc


def _es_monedero(rfc, emisores_con_complemento):
    """Un RFC cuenta como monedero si está en el padrón del SAT o si en este
    mismo lote emitió un complemento de combustible. Lo segundo importa para
    los clientes que solo tienen CSV: su emisor se reconoce porque otro
    cliente sí trae su CFDI."""
    if rfc in emisores_con_complemento:
        return True
    try:
        import monederos
        return rfc in monederos._PADRON_POR_RFC
    except Exception:                                   # pragma: no cover
        return False


def _promedio(total, meses):
    return round(total / meses, 2) if meses else None


def _cruzar_comisiones(fila, monederos_rfc, comisiones, cargado_por_mes):
    """Llena las columnas de comisión que vienen de Syntage.

    El porcentaje solo se calcula sobre los meses que tienen AMBAS cosas: la
    comisión facturada y el combustible cargado. Un mes con comisión pero sin
    XML descargado se reporta en el total histórico y se queda FUERA de la
    base — dividir la comisión de un mes entre el gasto de otro daría un
    porcentaje inventado.

    Las tres cifras son acumuladas (explícita ⊂ +administrativa ⊂ total)
    porque así se comparan contra un pricing: son tres respuestas a "cuánto
    paga hoy", según dónde se ponga la línea de qué es comisión."""
    rfc = fila["rfc_cliente"]
    desgloses = [comisiones[(rfc, m)] for m in monederos_rfc
                 if (rfc, m) in comisiones]
    if not desgloses:
        return

    fila["declara_cero"] = sum(d.get("declara_cero") or 0 for d in desgloses)

    # Los totales históricos que reporta el monedero, con el fondeo aparte:
    # el fondeo es el dinero movido, no un cargo, y es el denominador con el
    # que un monedero cobra.
    tot = {k: 0.0 for k in ("explicita", "administrativa", "otros", "fondeo",
                            "fondeo_combustible")}
    for d in desgloses:
        for k in tot:
            tot[k] += (d.get("total") or {}).get(k) or 0
    cargos = tot["explicita"] + tot["administrativa"] + tot["otros"]
    fila["fondeo"] = round(tot["fondeo"], 2)
    fila["fondeo_combustible"] = round(tot["fondeo_combustible"], 2)
    fila["cargos_monedero"] = round(cargos, 2)
    fila["comision_syntage_total"] = round(cargos, 2)
    if tot["fondeo"]:
        fila["pct_sobre_fondeo"] = cargos / tot["fondeo"]
        fila["pct_comision_sobre_fondeo"] = (
            (tot["explicita"] + tot["administrativa"]) / tot["fondeo"])
        fila["proporcion_combustible"] = tot["fondeo_combustible"] / tot["fondeo"]

    explicita = administrativa = otros = 0.0
    meses_pareados = set()
    for d in desgloses:
        for mes, montos in (d.get("meses") or {}).items():
            if (rfc, mes) not in cargado_por_mes:
                continue
            meses_pareados.add(mes)
            explicita += montos.get("explicita") or 0
            administrativa += montos.get("administrativa") or 0
            otros += montos.get("otros") or 0

    base = sum(cargado_por_mes[(rfc, m)] for m in meses_pareados)
    fila["meses_pareados"] = len(meses_pareados)
    fila["base_comision"] = round(base, 2)
    fila["syntage_explicita"] = round(explicita, 2)
    fila["syntage_administrativa"] = round(administrativa, 2)
    fila["syntage_otros"] = round(otros, 2)
    if base:
        fila["pct_explicita"] = explicita / base
        fila["pct_administrativa"] = (explicita + administrativa) / base
        fila["pct_total"] = (explicita + administrativa + otros) / base


def analizar(ruta, comisiones=None):
    """Lee todas las facturas de `ruta` (un zip o una carpeta) y devuelve los
    agregados del discovery.

    `comisiones` es opcional: el resultado de
    `comisiones_monedero.recolectar()`, indexado por (rfc_cliente,
    rfc_monedero). Sin él, este módulo no toca la red y las columnas de
    comisión de Syntage quedan en None — que es distinto de cero."""
    comisiones = comisiones or {}
    cfdis = {}
    filas_csv = {}
    for fuente in _fuentes(ruta):
        nombre, ext, datos = fuente
        if ext == ".csv":
            for f in _leer_csv(datos):
                filas_csv.setdefault(f["folio_fiscal"], f)
            continue
        registro = _leer(fuente)
        if registro is None:
            continue
        # Deduplicar por folio fiscal: el mismo CFDI llega en varias descargas.
        cfdis.setdefault(registro["folio_fiscal"], registro)

    sospechosos = []
    # (rfc_cliente, mes) -> importe cargado, para cruzar contra la comisión.
    cargado_por_mes = {}
    clientes = {}
    estaciones = {}
    otros_cargos = []
    emisores_con_complemento = set()

    # Primera pasada: qué emisores son monedero de verdad (emitieron un
    # complemento) y qué CFDI no se pueden usar. Se separa de la suma porque
    # un descuadre se descubre a media factura y deshacer sumas ya hechas es
    # justo la clase de error que este módulo debe evitar.
    def _fuera(reg, motivo, declarado=None, sumado=None):
        sospechosos.append({
            "folio_fiscal": reg["folio_fiscal"],
            "rfc_cliente": reg["rfc_cliente"],
            "rfc_monedero": reg["rfc_monedero"],
            "mes": reg["mes"],
            "motivo": motivo,
            "declarado": declarado,
            "sumado": sumado,
            "archivo": reg["archivo"],
        })

    for reg in cfdis.values():
        # Sin RFC de cliente o de monedero la factura no se puede atribuir, y
        # sus cargos inflarían un ranking sin dueño. Se marca y se deja fuera.
        if not reg["rfc_cliente"] or not reg["rfc_monedero"]:
            _fuera(reg, "sin RFC de cliente o de monedero")
            continue
        if reg["resumen"] is None:
            continue
        emisores_con_complemento.add(reg["rfc_monedero"])
        if not ecm.cuadra(reg["cargos"], reg["resumen"]):
            _fuera(reg, "el subtotal declarado no cuadra con la suma de cargos",
                   declarado=reg["resumen"]["subtotal"],
                   sumado=round(sum(c["importe"] or 0 for c in reg["cargos"]), 2))

    folios_sospechosos = {s["folio_fiscal"] for s in sospechosos}

    for reg in cfdis.values():
        if reg["folio_fiscal"] in folios_sospechosos:
            continue
        rfc_cliente = reg["rfc_cliente"]
        rfc_monedero = reg["rfc_monedero"]
        c = clientes.setdefault(rfc_cliente, {
            "rfc_cliente": rfc_cliente,
            "razon_social": reg["razon_social"],
            "monederos_rfc": set(),
            "meses_set": set(),
            "litros": 0.0,
            "importe": 0.0,
            "cargas": 0,
            "comision": 0.0,
            "comision_por_mes": {},
        })
        c["monederos_rfc"].add(rfc_monedero)
        if reg["razon_social"] and not c["razon_social"]:
            c["razon_social"] = reg["razon_social"]

        if reg["resumen"] is not None:
            c["meses_set"].add(reg["mes"])
            for cargo in reg["cargos"]:
                importe = cargo["importe"] or 0
                litros = cargo["cantidad"] or 0
                c["importe"] += importe
                c["litros"] += litros
                c["cargas"] += 1
                # La unidad es el RFC de la estación, no el par (RFC, clave).
                # Se confirmó con datos reales: la clave viene en "0" en 244
                # de 397 casos y un mismo RFC llega a tener 18 claves. Partir
                # por clave rompe la estación más grande en pedazos y la hace
                # ver como si la usara un solo cliente — justo al contrario de
                # lo que sirve para decidir a qué estación acercarse.
                e = estaciones.setdefault(cargo["rfc_estacion"], {
                    "rfc_estacion": cargo["rfc_estacion"],
                    "claves_set": set(),
                    "cargas": 0,
                    "litros": 0.0,
                    "importe": 0.0,
                    "clientes_set": set(),
                    "meses_set": set(),
                    "monederos_set": set(),
                })
                if cargo["clave_estacion"]:
                    e["claves_set"].add(cargo["clave_estacion"])
                e["cargas"] += 1
                e["litros"] += litros
                e["importe"] += importe
                e["clientes_set"].add(rfc_cliente)
                e["meses_set"].add(reg["mes"])
                e["monederos_set"].add(rfc_monedero)
            cargado_por_mes[(rfc_cliente, reg["mes"])] = (
                cargado_por_mes.get((rfc_cliente, reg["mes"]), 0.0)
                + (reg["resumen"]["subtotal"] or 0))
        else:
            if reg["comision"]:
                c["comision"] += reg["comision"]
                c["comision_por_mes"][reg["mes"]] = (
                    c["comision_por_mes"].get(reg["mes"], 0.0) + reg["comision"])
            for o in reg["otros_cargos"]:
                otros_cargos.append({
                    "rfc_cliente": rfc_cliente,
                    "razon_social": c["razon_social"],
                    "monedero": _nombre_monedero(rfc_monedero, reg["nombre_monedero"]),
                    "rfc_monedero": rfc_monedero,
                    "mes": reg["mes"],
                    "descripcion": o["descripcion"],
                    "importe": o["importe"],
                    "folio_fiscal": reg["folio_fiscal"],
                })

    # Nombres de monedero, tomados del CFDI cuando el padrón no los tiene.
    nombres_cfdi = {}
    for reg in cfdis.values():
        nombres_cfdi.setdefault(reg["rfc_monedero"], reg["nombre_monedero"])
    for f in filas_csv.values():
        nombres_cfdi.setdefault(f["rfc_monedero"], f["nombre_monedero"])

    # ── Clientes ────────────────────────────────────────────────────────────
    filas_clientes = []
    for c in clientes.values():
        # El % de comisión solo se calcula sobre los meses que tienen AMBAS
        # cosas: la comisión facturada y el gasto cargado. Dividir la comisión
        # de un mes entre el gasto de otro daría un porcentaje inventado.
        meses_pareados = [m for m in c["comision_por_mes"]
                          if (c["rfc_cliente"], m) in cargado_por_mes]
        comision_pareada = sum(c["comision_por_mes"][m] for m in meses_pareados)
        base = sum(cargado_por_mes[(c["rfc_cliente"], m)] for m in meses_pareados)
        fila = {
            "rfc_cliente": c["rfc_cliente"],
            "razon_social": c["razon_social"],
            "monederos": sorted(_nombre_monedero(r, nombres_cfdi.get(r))
                                for r in c["monederos_rfc"]),
            "monederos_rfc": sorted(c["monederos_rfc"]),
            "meses": len(c["meses_set"]),
            "cargas": c["cargas"],
            "litros": round(c["litros"], 2),
            "importe": round(c["importe"], 2),
            "promedio_mensual": _promedio(c["importe"], len(c["meses_set"])),
            "comision": round(c["comision"], 2),
            "comision_porcentaje": (comision_pareada / base) if base else None,
            # Comisión de Syntage. En None, no en cero: no consultar a un
            # cliente no es lo mismo que consultarlo y no encontrarle nada.
            "meses_pareados": 0,
            "base_comision": None,
            "syntage_explicita": None,
            "syntage_administrativa": None,
            "syntage_otros": None,
            "pct_explicita": None,
            "pct_administrativa": None,
            "pct_total": None,
            "comision_syntage_total": None,
            "declara_cero": 0,
            "fondeo": None,
            "fondeo_combustible": None,
            "cargos_monedero": None,
            "pct_sobre_fondeo": None,
            "pct_comision_sobre_fondeo": None,
            "proporcion_combustible": None,
        }
        _cruzar_comisiones(fila, c["monederos_rfc"], comisiones, cargado_por_mes)
        filas_clientes.append(fila)
    # Un cliente sin ningún complemento no es un cliente de este ranking:
    # aparece en `sin_detalle`, con lo poco que sí se sabe de él.
    con_detalle = {f["rfc_cliente"] for f in filas_clientes if f["meses"]}
    filas_clientes = [f for f in filas_clientes if f["meses"]]
    filas_clientes.sort(key=lambda f: -f["importe"])

    # ── Estaciones ──────────────────────────────────────────────────────────
    filas_estaciones = []
    for e in estaciones.values():
        filas_estaciones.append({
            "rfc_estacion": e["rfc_estacion"],
            "sucursales": len(e["claves_set"]),
            "claves": sorted(e["claves_set"]),
            "cargas": e["cargas"],
            "litros": round(e["litros"], 2),
            "importe": round(e["importe"], 2),
            "meses_activos": len(e["meses_set"]),
            "promedio_mensual": _promedio(e["importe"], len(e["meses_set"])),
            "clientes_distintos": len(e["clientes_set"]),
            "clientes": sorted(clientes[r]["razon_social"] or r
                               for r in e["clientes_set"]),
            "monederos": sorted(_nombre_monedero(r, nombres_cfdi.get(r))
                                for r in e["monederos_set"]),
        })
    # Manda el importe: el volumen que se puede redirigir a esa estación es el
    # premio. Los clientes distintos desempatan — entre dos estaciones del
    # mismo tamaño, la que ya sirve a varios se cierra una vez y sirve a todos.
    filas_estaciones.sort(key=lambda f: (-f["importe"], -f["clientes_distintos"]))

    # ── Monederos ───────────────────────────────────────────────────────────
    mon = {}
    for c in clientes.values():
        for rfc in c["monederos_rfc"]:
            m = mon.setdefault(rfc, {"clientes_set": set(), "importe": 0.0,
                                     "comision": 0.0, "base": 0.0})
            m["clientes_set"].add(c["rfc_cliente"])
    for reg in cfdis.values():
        if reg["folio_fiscal"] in folios_sospechosos:
            continue
        m = mon.setdefault(reg["rfc_monedero"], {"clientes_set": set(), "importe": 0.0,
                                                 "comision": 0.0, "base": 0.0})
        if reg["resumen"] is not None:
            m["importe"] += sum(c["importe"] or 0 for c in reg["cargos"])
        elif reg["comision"]:
            # Solo cuenta la comisión de un mes que también tenga gasto
            # cargado de ese mismo cliente: si no, no hay base contra la cual
            # sacar un porcentaje.
            base = cargado_por_mes.get((reg["rfc_cliente"], reg["mes"]))
            if base:
                m["comision"] += reg["comision"]
                m["base"] += base
    for f in filas_csv.values():
        if _es_monedero(f["rfc_monedero"], emisores_con_complemento):
            m = mon.setdefault(f["rfc_monedero"], {"clientes_set": set(), "importe": 0.0,
                                                   "comision": 0.0, "base": 0.0})
            m["clientes_set"].add(f["rfc_cliente"])

    filas_monederos = [{
        "rfc_monedero": rfc,
        "nombre": _nombre_monedero(rfc, nombres_cfdi.get(rfc)),
        "clientes": len(m["clientes_set"]),
        "importe": round(m["importe"], 2),
        "comision": round(m["comision"], 2),
        "comision_porcentaje": (m["comision"] / m["base"]) if m["base"] else None,
    } for rfc, m in mon.items()]
    filas_monederos.sort(key=lambda f: (-f["importe"], -f["clientes"]))

    # ── Sin detalle de estación ─────────────────────────────────────────────
    # Clientes de los que se sabe qué monedero usan (por CSV) pero de los que
    # no se bajó ningún CFDI con complemento. No son ceros: son la lista de
    # qué falta descargar.
    sin_detalle = {}
    for f in filas_csv.values():
        if not f["rfc_cliente"] or f["rfc_cliente"] in con_detalle:
            continue
        s = sin_detalle.setdefault(f["rfc_cliente"], {
            "rfc_cliente": f["rfc_cliente"],
            "razon_social": f["razon_social"],
            "monederos_rfc": set(),
            "facturas": 0,
            "meses_set": set(),
        })
        s["facturas"] += 1
        if f["mes"]:
            s["meses_set"].add(f["mes"])
        if _es_monedero(f["rfc_monedero"], emisores_con_complemento):
            s["monederos_rfc"].add(f["rfc_monedero"])
        if f["razon_social"] and not s["razon_social"]:
            s["razon_social"] = f["razon_social"]

    filas_sin_detalle = [{
        "rfc_cliente": s["rfc_cliente"],
        "razon_social": s["razon_social"],
        "monederos": sorted(_nombre_monedero(r, nombres_cfdi.get(r))
                            for r in s["monederos_rfc"]),
        "monederos_rfc": sorted(s["monederos_rfc"]),
        "facturas": s["facturas"],
        "meses": len(s["meses_set"]),
    } for s in sin_detalle.values()]
    filas_sin_detalle.sort(key=lambda f: -f["facturas"])

    otros_cargos.sort(key=lambda o: -o["importe"])

    return {
        "clientes": filas_clientes,
        "estaciones": filas_estaciones,
        "monederos": filas_monederos,
        "sin_detalle": filas_sin_detalle,
        "otros_cargos": otros_cargos,
        "sospechosos": sospechosos,
        "cfdis_leidos": len(cfdis),
        "facturas_en_csv": len(filas_csv),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────
def _pct(v):
    return None if v is None else round(v * 100, 2)


HOJAS = [
    ("Clientes", [
        ("RFC", lambda f: f["rfc_cliente"]),
        ("Razón social", lambda f: f["razon_social"]),
        ("Monedero(s)", lambda f: ", ".join(f["monederos"])),
        ("Meses con detalle", lambda f: f["meses"]),
        ("Cargas", lambda f: f["cargas"]),
        ("Litros", lambda f: f["litros"]),
        ("Importe total", lambda f: f["importe"]),
        ("Promedio mensual", lambda f: f["promedio_mensual"]),
        # Las tres cifras de comisión, acumuladas, según dónde se ponga la
        # línea de qué cuenta como comisión. Se reportan por separado porque
        # esa línea es una decisión de negociación, no técnica, y mueve el
        # resultado casi 4×.
        # Sobre el FONDEO: así cobra un monedero, un % del dinero que mueve.
        # Es la cifra comparable contra un pricing propio.
        ("Fondeo total $", lambda f: f["fondeo"]),
        ("Fondeo de combustible $", lambda f: f["fondeo_combustible"]),
        ("Del fondeo, es combustible %", lambda f: _pct(f["proporcion_combustible"])),
        ("Cargos del monedero $", lambda f: f["cargos_monedero"]),
        ("Comisión s/fondeo %", lambda f: _pct(f["pct_comision_sobre_fondeo"])),
        ("Todos los cargos s/fondeo %", lambda f: _pct(f["pct_sobre_fondeo"])),
        # Sobre el combustible MEDIDO en los XML. Solo vale cuando el
        # monedero no factura el fondeo (entonces no hay otro denominador),
        # y solo sobre los meses que tienen ambas cosas.
        ("Meses con comisión y combustible", lambda f: f["meses_pareados"]),
        ("Combustible en esos meses $", lambda f: f["base_comision"]),
        ("Comisión explícita %", lambda f: _pct(f["pct_explicita"])),
        ("Comisión + administrativa %", lambda f: _pct(f["pct_administrativa"])),
        ("Costo total monedero %", lambda f: _pct(f["pct_total"])),
        ("Veces que declaró no cobrar comisión", lambda f: f["declara_cero"]),
        ("Comisión vista en los CFDI $", lambda f: f["comision"]),
    ]),
    ("Estaciones", [
        ("RFC estación", lambda f: f["rfc_estacion"]),
        ("Importe total", lambda f: f["importe"]),
        ("Promedio mensual", lambda f: f["promedio_mensual"]),
        ("Clientes distintos", lambda f: f["clientes_distintos"]),
        ("Sucursales", lambda f: f["sucursales"]),
        ("Cargas", lambda f: f["cargas"]),
        ("Litros", lambda f: f["litros"]),
        ("Meses activos", lambda f: f["meses_activos"]),
        ("Clientes", lambda f: ", ".join(f["clientes"])),
        ("Monedero(s)", lambda f: ", ".join(f["monederos"])),
        ("Claves de estación", lambda f: ", ".join(f["claves"])),
    ]),
    ("Monederos", [
        ("RFC", lambda f: f["rfc_monedero"]),
        ("Monedero", lambda f: f["nombre"]),
        ("Clientes", lambda f: f["clientes"]),
        ("Importe total", lambda f: f["importe"]),
        ("Comisión $", lambda f: f["comision"]),
        ("Comisión %", lambda f: _pct(f["comision_porcentaje"])),
    ]),
    ("Sin detalle de estación", [
        ("RFC", lambda f: f["rfc_cliente"]),
        ("Razón social", lambda f: f["razon_social"]),
        ("Monedero(s)", lambda f: ", ".join(f["monederos"])),
        ("Facturas en CSV", lambda f: f["facturas"]),
        ("Meses", lambda f: f["meses"]),
    ]),
    ("Otros cargos", [
        ("RFC cliente", lambda f: f["rfc_cliente"]),
        ("Razón social", lambda f: f["razon_social"]),
        ("Monedero", lambda f: f["monedero"]),
        ("Mes", lambda f: f["mes"]),
        ("Descripción", lambda f: f["descripcion"]),
        ("Importe", lambda f: f["importe"]),
        ("Folio fiscal", lambda f: f["folio_fiscal"]),
    ]),
    ("Sospechosos", [
        ("RFC cliente", lambda f: f["rfc_cliente"]),
        ("RFC monedero", lambda f: f["rfc_monedero"]),
        ("Mes", lambda f: f["mes"]),
        ("Motivo", lambda f: f["motivo"]),
        ("Subtotal declarado", lambda f: f["declarado"]),
        ("Suma de cargos", lambda f: f["sumado"]),
        ("Folio fiscal", lambda f: f["folio_fiscal"]),
        ("Archivo", lambda f: f["archivo"]),
    ]),
]

_LLAVE_DE_HOJA = {
    "Clientes": "clientes",
    "Estaciones": "estaciones",
    "Monederos": "monederos",
    "Sin detalle de estación": "sin_detalle",
    "Otros cargos": "otros_cargos",
    "Sospechosos": "sospechosos",
}


def escribir_xlsx(resultado, destino):
    """Las seis hojas del discovery en un solo archivo. Se importa openpyxl
    aquí y no arriba para que `analizar()` siga sirviendo sin tenerlo
    instalado — mismo criterio que el resto del repo con las dependencias
    opcionales."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    libro = Workbook()
    libro.remove(libro.active)
    encabezado_fondo = PatternFill("solid", fgColor="F1654B")
    encabezado_letra = Font(bold=True, color="FFFFFF")

    for titulo, columnas in HOJAS:
        hoja = libro.create_sheet(titulo)
        for i, (nombre, _) in enumerate(columnas, start=1):
            celda = hoja.cell(row=1, column=i, value=nombre)
            celda.fill = encabezado_fondo
            celda.font = encabezado_letra
            celda.alignment = Alignment(vertical="center", wrap_text=True)
        for j, fila in enumerate(resultado[_LLAVE_DE_HOJA[titulo]], start=2):
            for i, (_, saca) in enumerate(columnas, start=1):
                hoja.cell(row=j, column=i, value=saca(fila))
        hoja.freeze_panes = "A2"
        for i, (nombre, _) in enumerate(columnas, start=1):
            ancho = max(len(nombre) + 2, 12)
            if nombre in ("Clientes", "Razón social", "Monedero(s)",
                          "Descripción", "Archivo"):
                ancho = 42
            hoja.column_dimensions[hoja.cell(row=1, column=i).column_letter].width = ancho

    carpeta = os.path.dirname(os.path.abspath(destino))
    if carpeta:
        os.makedirs(carpeta, exist_ok=True)
    libro.save(destino)
    return destino


# ─────────────────────────────────────────────────────────────────────────────
def _pesos(v):
    return "$%s" % format(v or 0, ",.2f")


def _porciento(v):
    return "%.2f%%" % (v * 100) if v is not None else "—"


def _tabla_comisiones(r, tope):
    # El porcentaje va sobre el FONDEO —el dinero que pasa por el monedero—
    # porque así cobra un monedero. Sobre el combustible medido en los XML
    # daría porcentajes absurdos: hay clientes que mueven $4.5M de despensa y
    # restaurante contra $73 mil de gasolina con el mismo monedero.
    con_fondeo = [c for c in r["clientes"] if c.get("pct_sobre_fondeo") is not None]
    if con_fondeo:
        print("\nCOMISIÓN sobre el fondeo — el dinero que pasa por el monedero")
        print("  %-30s %15s %10s %11s %11s" % (
            "Cliente", "Fondeo", "% combust.", "Comisión", "Con cargos"))
        for c in sorted(con_fondeo, key=lambda f: -(f["fondeo"] or 0)):
            print("  %-30s %15s %10s %11s %11s" % (
                (c["razon_social"] or "")[:30], _pesos(c["fondeo"]),
                _porciento(c["proporcion_combustible"]),
                _porciento(c["pct_comision_sobre_fondeo"]),
                _porciento(c["pct_sobre_fondeo"])))
        print("\n  '%% combust.' es qué parte del fondeo es gasolina. Donde es "
              "bajo, la comisión\n  es en buena medida de otro producto "
              "(despensa, restaurante), no del combustible.")

    sin_fondeo = [c for c in r["clientes"]
                  if c.get("pct_sobre_fondeo") is None and c["meses_pareados"]]
    if sin_fondeo:
        print("\n  Su monedero NO factura el fondeo, así que el único "
              "denominador es el\n  combustible medido en los XML (solo los "
              "meses que tienen ambas cosas):")
        print("  %-30s %6s %14s %11s %11s %11s" % (
            "Cliente", "Meses", "Combustible", "Explícita", "+ Admin.", "Todo"))
        for c in sin_fondeo:
            print("  %-30s %6d %14s %11s %11s %11s" % (
                (c["razon_social"] or "")[:30], c["meses_pareados"],
                _pesos(c["base_comision"]), _porciento(c["pct_explicita"]),
                _porciento(c["pct_administrativa"]), _porciento(c["pct_total"])))

    sin_base = [c for c in r["clientes"] + r["sin_detalle"]
                if not c.get("meses_pareados") and c.get("cargos_monedero")
                and c.get("pct_sobre_fondeo") is None]
    if sin_base:
        print("\n  Le facturan cargos, pero no hay fondeo ni combustible del "
              "mismo mes para sacar un porcentaje:")
        for c in sin_base[:tope]:
            print("  %-30s %14s en cargos de su monedero" % (
                (c["razon_social"] or c["rfc_cliente"])[:30],
                _pesos(c["cargos_monedero"])))

    declara = [c for c in r["clientes"] if c.get("declara_cero")]
    if declara:
        print("\n  Facturas donde el monedero declara explícitamente que NO "
              "cobra comisión (dato de negociación, no un cero):")
        for c in declara:
            print("  %-30s %d vez(ces)" % (
                (c["razon_social"] or "")[:30], c["declara_cero"]))


def imprimir_resumen(r, tope=12, con_comisiones=False):
    print("%d CFDI leídos, %d facturas listadas en CSV.\n"
          % (r["cfdis_leidos"], r["facturas_en_csv"]))

    print("CLIENTES por gasto en combustible (%d con detalle)" % len(r["clientes"]))
    print("  %-14s %-30s %-24s %6s %14s %9s" % (
        "RFC", "Razón social", "Monedero", "Meses", "Importe", "Comisión"))
    for c in r["clientes"][:tope]:
        pct = ("%.2f%%" % (c["comision_porcentaje"] * 100)
               if c["comision_porcentaje"] is not None else "—")
        print("  %-14s %-30s %-24s %6d %14s %9s" % (
            c["rfc_cliente"], (c["razon_social"] or "")[:30],
            (", ".join(c["monederos"]))[:24], c["meses"],
            _pesos(c["importe"]), pct))

    con_varios = [e for e in r["estaciones"] if e["clientes_distintos"] > 1]
    print("\nESTACIONES por importe (%d RFC distintos, %d con más de un cliente)"
          % (len(r["estaciones"]), len(con_varios)))
    print("  %-14s %14s %13s %9s %11s %7s" % (
        "RFC estación", "Importe", "Prom./mes", "Clientes", "Sucursales", "Cargas"))
    for e in r["estaciones"][:tope]:
        print("  %-14s %14s %13s %9d %11d %7d" % (
            e["rfc_estacion"], _pesos(e["importe"]), _pesos(e["promedio_mensual"]),
            e["clientes_distintos"], e["sucursales"], e["cargas"]))

    if con_varios:
        print("\n  Las que ya sirven a VARIOS clientes tuyos — se cierran una vez "
              "y sirven a todos:")
        for e in sorted(con_varios,
                        key=lambda f: (-f["clientes_distintos"], -f["importe"]))[:tope]:
            print("  %-14s %14s  %d clientes: %s" % (
                e["rfc_estacion"], _pesos(e["importe"]), e["clientes_distintos"],
                (", ".join(e["clientes"]))[:60]))

    print("\nMONEDEROS")
    print("  %-14s %-34s %9s %14s %10s" % (
        "RFC", "Monedero", "Clientes", "Importe", "Comisión"))
    for m in r["monederos"]:
        pct = ("%.2f%%" % (m["comision_porcentaje"] * 100)
               if m["comision_porcentaje"] is not None else "—")
        print("  %-14s %-34s %9d %14s %10s" % (
            m["rfc_monedero"], m["nombre"][:34], m["clientes"],
            _pesos(m["importe"]), pct))

    if con_comisiones:
        _tabla_comisiones(r, tope)
    else:
        print("\n  (La comisión no está en los archivos descargados: son estados "
              "de cuenta.\n   Se lee de Syntage agregando --comisiones.)")

    if r["sin_detalle"]:
        print("\nSIN DETALLE DE ESTACIÓN — se sabe el monedero, falta bajar el CFDI")
        for s in r["sin_detalle"]:
            print("  %-14s %-30s %-30s %d factura(s) en CSV" % (
                s["rfc_cliente"], (s["razon_social"] or "")[:30],
                (", ".join(s["monederos"]) or "monedero no identificado")[:30],
                s["facturas"]))

    if r["sospechosos"]:
        print("\nATENCIÓN: %d CFDI no cuadraron y quedaron FUERA del análisis"
              % len(r["sospechosos"]))
        for s in r["sospechosos"]:
            detalle = ""
            if s["declarado"] is not None:
                detalle = " (declarado %s vs sumado %s)" % (
                    _pesos(s["declarado"]), _pesos(s["sumado"]))
            print("  %-14s %-8s %s%s" % (
                s["rfc_cliente"] or "sin RFC", s["mes"] or "", s["motivo"], detalle))

    if r["otros_cargos"]:
        total = sum(o["importe"] for o in r["otros_cargos"])
        print("\n%d cargo(s) que NO son comisión, por %s — se reportan aparte:"
              % (len(r["otros_cargos"]), _pesos(total)))
        for o in r["otros_cargos"][:tope]:
            print("  %-14s %-24s %-40s %10s" % (
                o["rfc_cliente"], o["monedero"][:24], o["descripcion"][:40],
                _pesos(o["importe"])))


def pares_cliente_monedero(r):
    """Los (rfc_cliente, rfc_monedero) que hay que consultarle a Syntage,
    sacados de un análisis ya hecho. Va por RFC y no por nombre comercial:
    un monedero puede aparecer con dos nombres (el del padrón y el del CFDI)
    y la API se consulta por RFC.

    Incluye a los clientes sin detalle de estación: de ellos no se puede
    sacar un porcentaje —falta la base— pero sí cuánto les factura su
    monedero, que ya es dato de negociación."""
    pares = set()
    for c in r["clientes"] + r["sin_detalle"]:
        for rfc_monedero in c["monederos_rfc"]:
            pares.add((c["rfc_cliente"], rfc_monedero))
    return pares


def main(argv):
    if len(argv) < 3 or argv[1] not in ("resumen", "xlsx"):
        print(__doc__.split("Uso:")[1].split("**")[0].strip())
        return 1

    ruta = argv[2]
    if not os.path.exists(ruta):
        print("No existe: %s" % ruta)
        return 1

    con_comisiones = "--comisiones" in argv
    r = analizar(ruta)

    if con_comisiones:
        # La comisión no está en los archivos descargados: se lee de Syntage.
        # Se necesita el análisis offline primero para saber a qué pares
        # preguntarle, y por eso se analiza dos veces — leer 86 XML es
        # barato, y así este módulo no decide solo cuándo tocar la red.
        import comisiones_monedero
        pares = pares_cliente_monedero(r)
        print("Consultando la comisión de %d (cliente, monedero) en Syntage. "
              "La primera vez tarda; después sale del caché en out/comisiones.\n"
              % len(pares))
        comisiones = comisiones_monedero.recolectar(
            sorted(pares), aviso=lambda t: print("  %s" % t))
        print()
        r = analizar(ruta, comisiones=comisiones)

    imprimir_resumen(r, con_comisiones=con_comisiones)

    if argv[1] == "xlsx":
        # Los argumentos que empiezan con "--" son banderas, no rutas: sin
        # este filtro, `xlsx <zip> --comisiones` guardaba el Excel en un
        # archivo llamado "--comisiones".
        sueltos = [a for a in argv[3:] if not a.startswith("--")]
        destino = sueltos[0] if sueltos else os.path.join(
            RAIZ, "out", "discovery_estaciones.xlsx")
        escribir_xlsx(r, destino)
        print("\nExcel: %s" % destino)
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main(sys.argv))
